from __future__ import annotations

import csv
import math
import random
import statistics
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook


N = 100
DEFAULT_REFERENCE_NAME = "0418"
RANDOM_SEED = None
LAB_ROOT = Path("lab-generator")
DATA_DIR = LAB_ROOT / "data"
OUTPUT_DIR = LAB_ROOT / "output"
MAX_ATTEMPTS_MULTIPLIER = 200
D65_VECTOR_ATTEMPTS = 50
BOUNDARY_WARNING_RATIO = 0.35


UV_ALIASES = {
    "UVAdj": "0%UV",
    "100% 完全": "100%UV",
}

REQUIRED_COLUMNS = {
    "uv": "UV 設定",
    "f2_l": "L*(10°/F2)",
    "f2_a": "a*(10°/F2)",
    "f2_b": "b*(10°/F2)",
    "d65_l": "L*(10°/D65)",
    "d65_a": "a*(10°/D65)",
    "d65_b": "b*(10°/D65)",
}

GENERATED_COLUMNS = [
    "No.",
    "0%UV_F2_L",
    "0%UV_F2_a",
    "0%UV_F2_b",
    "0%UV_D65_L",
    "0%UV_D65_a",
    "0%UV_D65_b",
    "100%UV_F2_L",
    "100%UV_F2_a",
    "100%UV_F2_b",
    "100%UV_D65_L",
    "100%UV_D65_a",
    "100%UV_D65_b",
]


@dataclass(frozen=True)
class LabRow:
    f2_l: float
    f2_a: float
    f2_b: float
    d65_l: float
    d65_a: float
    d65_b: float

    @property
    def vector(self) -> tuple[float, float, float]:
        return (
            self.d65_l - self.f2_l,
            self.d65_a - self.f2_a,
            self.d65_b - self.f2_b,
        )


@dataclass(frozen=True)
class UvModel:
    label: str
    rows: list[LabRow]
    f2_ranges: tuple[tuple[float, float], tuple[float, float], tuple[float, float]]
    d65_ranges: tuple[tuple[float, float], tuple[float, float], tuple[float, float]]
    vectors: list[tuple[float, float, float]]
    vector_std: tuple[float, float, float]


def main() -> int:
    if RANDOM_SEED is not None:
        random.seed(RANDOM_SEED)

    reference_name = get_reference_name()
    reference_csv = DATA_DIR / f"{reference_name}.csv"
    output_xlsx = OUTPUT_DIR / f"{reference_name}_generated_lab_output.xlsx"

    if not reference_csv.exists():
        print(
            f"找不到參考 CSV：{reference_csv}\n"
            "請將參考檔案放到上述路徑後再執行；程式不會自行捏造 CSV。"
        )
        return 1

    rows = read_reference_rows(reference_csv)
    models = build_models(rows)
    generated_rows = generate_rows(models, N)
    boundary_rows = build_boundary_check(generated_rows, models)
    duplicate_report = build_f2_duplicate_report(generated_rows, N)
    validate_generated_rows(generated_rows, models, duplicate_report)
    write_boundary_warnings(boundary_rows)
    write_excel(generated_rows, models, boundary_rows, output_xlsx)

    print(f"成功讀取參考檔：{reference_csv}")
    print(f"成功產生 {N} 組資料：{output_xlsx}")
    print_validation_report(duplicate_report)
    return 0


def get_reference_name() -> str:
    if len(sys.argv) >= 2:
        return Path(sys.argv[1]).stem
    return DEFAULT_REFERENCE_NAME


def read_reference_rows(path: Path) -> dict[str, list[LabRow]]:
    grouped: dict[str, list[LabRow]] = {"0%UV": [], "100%UV": []}

    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        if reader.fieldnames is None:
            raise ValueError(f"CSV 沒有標題列：{path}")

        normalized_columns = map_required_columns(reader.fieldnames)
        for row_number, row in enumerate(reader, start=2):
            uv_value = (row.get(normalized_columns["uv"]) or "").strip()
            uv_label = UV_ALIASES.get(uv_value)
            if uv_label is None:
                continue

            try:
                grouped[uv_label].append(
                    LabRow(
                        f2_l=parse_float(row[normalized_columns["f2_l"]]),
                        f2_a=parse_float(row[normalized_columns["f2_a"]]),
                        f2_b=parse_float(row[normalized_columns["f2_b"]]),
                        d65_l=parse_float(row[normalized_columns["d65_l"]]),
                        d65_a=parse_float(row[normalized_columns["d65_a"]]),
                        d65_b=parse_float(row[normalized_columns["d65_b"]]),
                    )
                )
            except ValueError as exc:
                raise ValueError(f"第 {row_number} 列 Lab 數值無法解析：{exc}") from exc

    for uv_label, lab_rows in grouped.items():
        if not lab_rows:
            raise ValueError(f"參考 CSV 中找不到 {uv_label} 對應資料。")

    return grouped


def map_required_columns(fieldnames: list[str]) -> dict[str, str]:
    by_clean_name = {clean_column_name(name): name for name in fieldnames}
    mapped: dict[str, str] = {}
    missing: list[str] = []

    for key, expected_name in REQUIRED_COLUMNS.items():
        actual_name = by_clean_name.get(clean_column_name(expected_name))
        if actual_name is None:
            missing.append(expected_name)
        else:
            mapped[key] = actual_name

    if missing:
        raise ValueError("CSV 缺少必要欄位：" + ", ".join(missing))

    return mapped


def clean_column_name(name: str) -> str:
    return name.strip().replace("\ufeff", "")


def parse_float(value: str) -> float:
    cleaned = (value or "").strip()
    if cleaned in {"", "---"}:
        raise ValueError(f"空白或非數值欄位：{value!r}")
    return float(cleaned)


def build_models(rows_by_uv: dict[str, list[LabRow]]) -> dict[str, UvModel]:
    return {label: build_model(label, rows) for label, rows in rows_by_uv.items()}


def build_model(label: str, rows: list[LabRow]) -> UvModel:
    vectors = [row.vector for row in rows]
    return UvModel(
        label=label,
        rows=rows,
        f2_ranges=(
            min_max(row.f2_l for row in rows),
            min_max(row.f2_a for row in rows),
            min_max(row.f2_b for row in rows),
        ),
        d65_ranges=(
            min_max(row.d65_l for row in rows),
            min_max(row.d65_a for row in rows),
            min_max(row.d65_b for row in rows),
        ),
        vectors=vectors,
        vector_std=(
            safe_std(vector[0] for vector in vectors),
            safe_std(vector[1] for vector in vectors),
            safe_std(vector[2] for vector in vectors),
        ),
    )


def min_max(values: object) -> tuple[float, float]:
    value_list = list(values)
    return min(value_list), max(value_list)


def safe_std(values: object) -> float:
    value_list = list(values)
    if len(value_list) < 2:
        return 0.01

    std = statistics.stdev(value_list)
    return std if std > 0 else 0.01


def generate_rows(models: dict[str, UvModel], count: int) -> list[list[float | int]]:
    generated: list[list[float | int]] = []
    zero_uv_f2_counts: Counter[tuple[float, float, float]] = Counter()
    full_uv_f2_counts: Counter[tuple[float, float, float]] = Counter()
    allowed_occurrences = get_allowed_occurrences(count)
    attempts = 0
    max_attempts = max(count * MAX_ATTEMPTS_MULTIPLIER, 1_000)

    while len(generated) < count and attempts < max_attempts:
        attempts += 1
        zero_uv = generate_lab_pair(models["0%UV"])
        full_uv = generate_lab_pair(models["100%UV"])
        zero_uv_f2_key = zero_uv[:3]
        full_uv_f2_key = full_uv[:3]

        if zero_uv_f2_counts[zero_uv_f2_key] >= allowed_occurrences:
            continue
        if full_uv_f2_counts[full_uv_f2_key] >= allowed_occurrences:
            continue

        zero_uv_f2_counts[zero_uv_f2_key] += 1
        full_uv_f2_counts[full_uv_f2_key] += 1
        values = (*zero_uv, *full_uv)
        generated.append([len(generated) + 1, *values])

    if len(generated) < count:
        zero_max = max(zero_uv_f2_counts.values(), default=0)
        full_max = max(full_uv_f2_counts.values(), default=0)
        raise RuntimeError(
            f"無法在最大嘗試次數內符合 F2 重複容許規則。"
            f" 0%UV_F2 Lab 最大出現次數={zero_max}，"
            f"100%UV_F2 Lab 最大出現次數={full_max}，"
            f"N={count}，allowed_occurrences={allowed_occurrences}。"
            "請放寬條件、增加小數位數，或降低生成筆數。"
        )

    return generated


def get_allowed_occurrences(count: int) -> int:
    return math.floor(count / 100) + 1


def generate_lab_pair(model: UvModel) -> tuple[float, float, float, float, float, float]:
    f2 = tuple(round_clamped(random.uniform(low, high), (low, high)) for low, high in model.f2_ranges)
    d65 = generate_d65_from_f2(f2, model)
    rounded_d65 = tuple(
        round_clamped(d65[index], model.d65_ranges[index])
        for index in range(3)
    )
    return (*f2, *rounded_d65)


def generate_d65_from_f2(f2: tuple[float, float, float], model: UvModel) -> tuple[float, float, float]:
    last_candidate: tuple[float, float, float] | None = None

    for _ in range(D65_VECTOR_ATTEMPTS):
        sampled_vector = random.choice(model.vectors)
        noisy_vector = tuple(
            sampled_vector[index] + random.gauss(0, model.vector_std[index] * 0.25)
            for index in range(3)
        )
        candidate = tuple(f2[index] + noisy_vector[index] for index in range(3))
        last_candidate = candidate

        if is_within_ranges(candidate, model.d65_ranges):
            return candidate

    if last_candidate is None:
        raise RuntimeError(f"{model.label} 沒有可用的 F2→D65 向量。")

    return tuple(
        clamp(last_candidate[index], *model.d65_ranges[index])
        for index in range(3)
    )


def is_within_ranges(
    values: tuple[float, float, float],
    ranges: tuple[tuple[float, float], tuple[float, float], tuple[float, float]],
) -> bool:
    return all(low <= values[index] <= high for index, (low, high) in enumerate(ranges))


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def round_clamped(value: float, value_range: tuple[float, float]) -> float:
    low, high = value_range
    return round(clamp(round(value, 2), round(low, 2), round(high, 2)), 2)


def build_boundary_check(
    generated_rows: list[list[float | int]],
    models: dict[str, UvModel],
) -> list[list[str | float | int]]:
    boundary_rows: list[list[str | float | int]] = []
    count = len(generated_rows)
    ranges_by_column = build_generated_column_ranges(models)

    for index, column_name in enumerate(GENERATED_COLUMNS[1:], start=1):
        values = [float(row[index]) for row in generated_rows]
        min_value, max_value = ranges_by_column[column_name]
        rounded_min = round(min_value, 2)
        rounded_max = round(max_value, 2)
        min_count = sum(value == rounded_min for value in values)
        max_count = sum(value == rounded_max for value in values)
        min_ratio = min_count / count if count else 0
        max_ratio = max_count / count if count else 0
        warning = (
            "High boundary concentration"
            if min_ratio > BOUNDARY_WARNING_RATIO or max_ratio > BOUNDARY_WARNING_RATIO
            else "OK"
        )
        boundary_rows.append(
            [
                column_name,
                rounded_min,
                rounded_max,
                min_count,
                max_count,
                round(min_ratio, 4),
                round(max_ratio, 4),
                warning,
            ]
        )

    return boundary_rows


def build_generated_column_ranges(models: dict[str, UvModel]) -> dict[str, tuple[float, float]]:
    return {
        "0%UV_F2_L": models["0%UV"].f2_ranges[0],
        "0%UV_F2_a": models["0%UV"].f2_ranges[1],
        "0%UV_F2_b": models["0%UV"].f2_ranges[2],
        "0%UV_D65_L": models["0%UV"].d65_ranges[0],
        "0%UV_D65_a": models["0%UV"].d65_ranges[1],
        "0%UV_D65_b": models["0%UV"].d65_ranges[2],
        "100%UV_F2_L": models["100%UV"].f2_ranges[0],
        "100%UV_F2_a": models["100%UV"].f2_ranges[1],
        "100%UV_F2_b": models["100%UV"].f2_ranges[2],
        "100%UV_D65_L": models["100%UV"].d65_ranges[0],
        "100%UV_D65_a": models["100%UV"].d65_ranges[1],
        "100%UV_D65_b": models["100%UV"].d65_ranges[2],
    }


def build_f2_duplicate_report(generated_rows: list[list[float | int]], count: int) -> dict[str, int | bool]:
    zero_uv_counts = Counter(tuple(row[1:4]) for row in generated_rows)
    full_uv_counts = Counter(tuple(row[7:10]) for row in generated_rows)
    allowed_occurrences = get_allowed_occurrences(count)
    zero_uv_max = max(zero_uv_counts.values(), default=0)
    full_uv_max = max(full_uv_counts.values(), default=0)
    is_valid = zero_uv_max <= allowed_occurrences and full_uv_max <= allowed_occurrences
    return {
        "n": count,
        "allowed_occurrences": allowed_occurrences,
        "zero_uv_f2_max_occurrences": zero_uv_max,
        "full_uv_f2_max_occurrences": full_uv_max,
        "is_valid": is_valid,
    }


def validate_generated_rows(
    generated_rows: list[list[float | int]],
    models: dict[str, UvModel],
    duplicate_report: dict[str, int | bool],
) -> None:
    if not duplicate_report["is_valid"]:
        raise RuntimeError("生成結果不符合 F2 重複容許規則。")

    ranges_by_column = build_generated_column_ranges(models)
    for column_index, column_name in enumerate(GENERATED_COLUMNS[1:], start=1):
        low, high = ranges_by_column[column_name]
        rounded_low = round(low, 2)
        rounded_high = round(high, 2)
        for row in generated_rows:
            value = float(row[column_index])
            if value < rounded_low or value > rounded_high:
                raise RuntimeError(
                    f"{column_name} 數值 {value} 超出參考範圍 "
                    f"{rounded_low}~{rounded_high}。"
                )


def write_boundary_warnings(boundary_rows: list[list[str | float | int]]) -> None:
    warning_rows = [row for row in boundary_rows if row[-1] != "OK"]
    if not warning_rows:
        return

    print("Warning: boundary concentration detected")
    for row in warning_rows:
        print(f"- {row[0]}: min_ratio={row[5]}, max_ratio={row[6]}")


def print_validation_report(duplicate_report: dict[str, int | bool]) -> None:
    print(f"N 值：{duplicate_report['n']}")
    print(f"allowed_occurrences：{duplicate_report['allowed_occurrences']}")
    print(f"0%UV_F2 Lab 最大重複出現次數：{duplicate_report['zero_uv_f2_max_occurrences']}")
    print(f"100%UV_F2 Lab 最大重複出現次數：{duplicate_report['full_uv_f2_max_occurrences']}")
    print(f"是否符合新的 F2 重複容許規則：{duplicate_report['is_valid']}")


def write_excel(
    generated_rows: list[list[float | int]],
    models: dict[str, UvModel],
    boundary_rows: list[list[str | float | int]],
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    generated_sheet = workbook.active
    generated_sheet.title = "generated_data"
    generated_sheet.append(GENERATED_COLUMNS)
    for row in generated_rows:
        generated_sheet.append(row)

    vectors_sheet = workbook.create_sheet("conversion_vectors")
    vectors_sheet.append(["UV_condition", "No.", "D65_L_minus_F2_L", "D65_a_minus_F2_a", "D65_b_minus_F2_b"])
    for label in ("0%UV", "100%UV"):
        for index, vector in enumerate(models[label].vectors, start=1):
            vectors_sheet.append([label, index, *[round(value, 4) for value in vector]])

    boundary_sheet = workbook.create_sheet("boundary_check")
    boundary_sheet.append(
        [
            "column_name",
            "min_value",
            "max_value",
            "min_count",
            "max_count",
            "min_ratio",
            "max_ratio",
            "warning",
        ]
    )
    for row in boundary_rows:
        boundary_sheet.append(row)

    workbook.save(output_path)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"執行失敗：{exc}", file=sys.stderr)
        raise SystemExit(1)
